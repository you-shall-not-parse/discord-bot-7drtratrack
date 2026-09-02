from types import SimpleNamespace

from openpyxl import Workbook

from cogs.rollcall import BLUEBERRY_ROLE_ID, RollCallCog


def test_blueberry_role_is_treated_as_former_member() -> None:
    blueberry = SimpleNamespace(roles=[SimpleNamespace(id=BLUEBERRY_ROLE_ID)])
    active = SimpleNamespace(roles=[SimpleNamespace(id=123)])

    assert RollCallCog._is_former_member(None)
    assert RollCallCog._is_former_member(blueberry)
    assert not RollCallCog._is_former_member(active)


def test_blueberry_is_excluded_from_expected_rollcall_members() -> None:
    blueberry = SimpleNamespace(
        id=1,
        display_name="Virid",
        roles=[SimpleNamespace(id=BLUEBERRY_ROLE_ID)],
    )
    active = SimpleNamespace(
        id=2,
        display_name="Active Member",
        roles=[SimpleNamespace(id=999)],
    )
    tracked_role = SimpleNamespace(members=[blueberry, active])
    guild = SimpleNamespace(get_role=lambda role_id: tracked_role if role_id == 999 else None)
    config = SimpleNamespace(tracked_role_ids=(999,), tracked_role_id=None)
    cog = object.__new__(RollCallCog)

    assert cog._expected_members(guild, config) == [active]


def test_rollcall_stores_formula_like_nickname_as_text() -> None:
    worksheet = Workbook().active
    worksheet.append(["User ID", "Nickname"])
    cog = object.__new__(RollCallCog)

    row = cog._upsert_member_row(worksheet, 123, "=HYPERLINK(\"https://example.com\")")

    cell = worksheet.cell(row=row, column=2)
    assert cell.value == "'=HYPERLINK(\"https://example.com\")"
    assert cell.data_type == "s"


def test_rollcall_html_uses_website_background_and_mobile_table_scroll() -> None:
    worksheet = Workbook().active
    worksheet.append(["User ID", "Nickname", "W36 07/09/2026"])
    worksheet.append([123, "Active Member", "Present"])
    guild = SimpleNamespace(get_member=lambda member_id: SimpleNamespace(roles=[]))
    config = SimpleNamespace(title="Test Roll Call")
    cog = object.__new__(RollCallCog)

    rendered = cog._render_html(guild, worksheet, config, highlight_week="W36 07/09/2026")

    assert "https://hllfrontlines.com/assets/website_backg.webp?v=1" in rendered
    assert 'class="table-scroll"' in rendered
    assert "background-attachment: scroll" in rendered
