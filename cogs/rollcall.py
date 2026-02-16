import asyncio
import html
import io
import json
import logging
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from typing import Optional

import discord
from discord import app_commands
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from discord.ext import commands
from openpyxl import Workbook, load_workbook

from data_paths import data_path

logger = logging.getLogger(__name__)


# =============================
# CONFIG (EDIT THIS)
# =============================

# Target guild
GUILD_ID = 1097913605082579024

# When to send the rollcall each week.
# NOTE: Columns are based on the date the rollcall is sent.
TIMEZONE = "Europe/London"  # uses pytz tz database name
SCHEDULE_WEEKDAY = "mon"  # mon/tue/wed/thu/fri/sat/sun
SCHEDULE_HOUR = 7  # 24h format, local time
SCHEDULE_MINUTE = 0

# Emoji members should react with to mark attendance.
ROLLCALL_EMOJI = "✅"

# Lock roll calls after this many hours from when the roll call post is sent.
# After lock, ✅ reactions are removed and the user is DM'd.
# Set to 0 or None to disable locking.
ROLLCALL_LOCK_HOURS: Optional[float] = 144.0

# Where to store the attendance workbook.
# The bot writes to .xlsx. If you have a legacy .xls, see IMPORT_LEGACY_XLS_PATH.
WORKBOOK_PATH = data_path("rollcall.xlsx")

# Optional: path to a legacy .xls workbook to import once (only if WORKBOOK_PATH doesn't exist yet).
# Requires pandas + xlrd==1.2.0 installed.
IMPORT_LEGACY_XLS_PATH: Optional[str] = None

# Optional: post HTML table uploads to a single channel.
# If None, HTML is posted to each rollcall channel.
HTML_CHANNEL_ID: Optional[int] = 1098525492631572567

# Whether to regenerate HTML on reaction add/remove.
# If False, HTML will still update on the backstop refresh (BACKSTOP_REFRESH_MINUTES) and on non-reaction refreshes.
UPDATE_HTML_ON_REACTION = True

# Whether to upload/replace the rollcall workbook message on reaction-driven refreshes.
# This can be very noisy because it deletes and re-posts the XLSX message each time.
UPLOAD_WORKBOOK_ON_REACTION = False

# Debounce delay for reaction-driven refreshes (seconds).
REACTION_REFRESH_DEBOUNCE_SECONDS = 30.0

# Optional: where to upload the rollcall.xlsx file so users can download it.
# If None, falls back to HTML_CHANNEL_ID (if set) or the first configured rollcall channel.
WORKBOOK_UPLOAD_CHANNEL_ID: Optional[int] = 1098525492631572567

# Backstop refresh for embeds/html in case of missed reaction events
BACKSTOP_REFRESH_MINUTES = 1440

# State file: message IDs + last posted week so we can edit across restarts
STATE_PATH = data_path("rollcall_state.json")

# Role allowed to use /forcerollcall
FORCE_ROLLCALL_ROLE_ID = 1213495462632361994

# Excluded statuses (shown at bottom of HTML, and excluded from "expected" list)
# Fill these with your server's role IDs.
HOMEGUARD_ROLE_ID: Optional[int] = 1103762811491975218
AWOL_ROLE_ID: Optional[int] = 1439416251687637044


@dataclass(frozen=True)
class RollCallConfig:
	key: str
	title: str
	channel_id: int
	tracked_role_id: Optional[int] = None  # legacy: single tracked role
	tracked_role_ids: Optional[tuple[int, ...]] = None  # optional: one or more tracked roles (use up to 2)
	ping_role_id: Optional[int] = None  # legacy: single ping role
	ping_role_ids: Optional[tuple[int, ...]] = None  # optional: one or more roles to mention when posting
	embed_image_url: Optional[str] = None  # optional CDN image URL to display on the embed


# Configure one entry per rollcall channel.
ROLLCALLS: list[RollCallConfig] = [
	    RollCallConfig(
	 	key="22nd",
	 	title="22nd Weekly Roll Call",
	 	channel_id=1099591160017719329,  # set the roll call channel ID
	 	tracked_role_ids=(1098347242202611732, 1099615408518070313),  # set the TWO role IDs allowed/expected to tick
	 	ping_role_ids=(1098347242202611732, 1099615408518070313),  # optional: role(s) to mention when posting
	 	embed_image_url="https://cdn.discordapp.com/attachments/1098976074852999261/1449441912770662491/file_000000002214722fa789165cdd45bc9b.png?ex=69934979&is=6991f7f9&hm=d67c84035d6f0685c2b7f9993167e81dec599e0be364ae927a04061c0b1a5119",
	 ),
	 	RollCallConfig(
	 	key="1-5th",
	 	title="1-5th Weekly Roll Call",
	 	channel_id=1472727556523425952,  # set the roll call channel ID
		 	tracked_role_ids=(1259814883248177196,),  # role(s) allowed/expected to tick
	 	ping_role_ids=(1259814883248177196,),  # optional: role(s) to mention when posting
	 	embed_image_url="https://cdn.discordapp.com/attachments/1098976074852999261/1444515451727253544/file_00000000e5f871f488f94dd458b30c09.png?ex=69932999&is=6991d819&hm=ba814b4a530031279073ec3fd49f4a4c1e34276586553afaa33839b5fb0ff81d",
	 ),
	 	RollCallConfig(
	 	key="InfantrySchool",
	 	title="Infantry School Weekly Roll Call",
	 	channel_id=1098331677224345660,  # set the roll call channel ID
		 	tracked_role_ids=(1099596178141757542,),  # role(s) allowed/expected to tick
	 	ping_role_ids=(1099596178141757542,),  # optional: role(s) to mention when posting
	 	embed_image_url="https://cdn.discordapp.com/attachments/1237437502248452227/1472736090094960801/IMG_2754.png?ex=6993a7dd&is=6992565d&hm=fd843d3d077addff34b6575655415a31f2f191c4f8684c71370adf5c9a400d9e",
	 ),
	 	RollCallConfig(
	 	key="8th",
	 	title="8th Weekly Roll Call",
	 	channel_id=1098701359022346341,  # set the roll call channel ID
	 	tracked_role_ids=(1099105947932168212, 1103626508645453975),  # set the TWO role IDs allowed/expected to tick
	 	ping_role_ids=(1099105947932168212, 1103626508645453975),  # optional: role(s) to mention when posting
	 	embed_image_url="https://cdn.discordapp.com/attachments/1098976074852999261/1444676650653450411/file_000000005384720e8f124201b4e379a9.png?ex=699316fa&is=6991c57a&hm=8a80e05652ea88fa5150df80f0b08cbbb71a0b89db034be287eec6c7813472f3",
	 ),
]


class RollCallCog(commands.Cog):
	def __init__(self, bot: commands.Bot):
		self.bot = bot
		self._lock = asyncio.Lock()
		self._state = self._load_state()
		self._scheduler: Optional[AsyncIOScheduler] = None
		self._backstop_task: Optional[asyncio.Task] = None
		self._refresh_task: Optional[asyncio.Task] = None
		self._debounce_task: Optional[asyncio.Task] = None

		# Don't start the scheduler or create asyncio tasks in __init__.
		# APScheduler jobs and create_task() require a running event loop.

	async def cog_load(self) -> None:
		# Called by discord.py when the cog is loaded (event loop is running).
		self._start_scheduler()
		if BACKSTOP_REFRESH_MINUTES and BACKSTOP_REFRESH_MINUTES > 0:
			if self._backstop_task is None or self._backstop_task.done():
				self._backstop_task = asyncio.create_task(self._backstop_loop())

	def _user_can_force(self, interaction: discord.Interaction) -> bool:
		user = interaction.user
		if isinstance(user, discord.Member):
			return any(r.id == FORCE_ROLLCALL_ROLE_ID for r in user.roles)
		return False

	def _excluded_role_markers(self) -> list[tuple[int, str]]:
		pairs: list[tuple[int, str]] = []
		if HOMEGUARD_ROLE_ID:
			pairs.append((int(HOMEGUARD_ROLE_ID), "HG"))
		if AWOL_ROLE_ID:
			pairs.append((int(AWOL_ROLE_ID), "AWOL"))
		return pairs

	def _member_exclusion_markers(self, member: discord.Member) -> list[str]:
		role_ids = {r.id for r in member.roles}
		markers: list[str] = []
		for rid, marker in self._excluded_role_markers():
			if rid in role_ids:
				markers.append(marker)
		return markers

	def _is_member_excluded(self, member: discord.Member) -> bool:
		return bool(self._member_exclusion_markers(member))

	def _excluded_members(self, guild: discord.Guild) -> list[discord.Member]:
		members: dict[int, discord.Member] = {}
		for rid, _marker in self._excluded_role_markers():
			role = guild.get_role(rid)
			if not role:
				continue
			for m in role.members:
				members[m.id] = m
		return sorted(members.values(), key=lambda m: (m.display_name or "").lower())

	@app_commands.command(name="forcerollcall", description="Force-run all configured roll calls now.")
	@app_commands.guilds(discord.Object(id=GUILD_ID))
	@app_commands.guild_only()
	async def forcerollcall(self, interaction: discord.Interaction) -> None:
		if interaction.guild_id != GUILD_ID:
			await interaction.response.send_message("This command can only be used in the main guild.", ephemeral=True)
			return
		if not self._user_can_force(interaction):
			await interaction.response.send_message("You don't have permission to use this command.", ephemeral=True)
			return

		await interaction.response.defer(ephemeral=True, thinking=True)
		try:
			await self._send_rollcalls(reason="forced_command")
			week_label = self._week_label(self._rollcall_date_for_now())
			await interaction.followup.send(f"Forced roll call run complete for **{week_label}**.", ephemeral=True)
		except Exception:
			logger.exception("/forcerollcall failed")
			await interaction.followup.send("Force roll call failed; check logs.", ephemeral=True)

	def cog_unload(self):
		if self._scheduler:
			try:
				self._scheduler.shutdown(wait=False)
			except Exception:
				pass
		if self._backstop_task and not self._backstop_task.done():
			self._backstop_task.cancel()
		if self._refresh_task and not self._refresh_task.done():
			self._refresh_task.cancel()
		if self._debounce_task and not self._debounce_task.done():
			self._debounce_task.cancel()

	# -----------------
	# State
	# -----------------
	def _load_state(self) -> dict:
		try:
			if not os.path.exists(STATE_PATH):
				return {"version": 1, "rollcalls": {}, "workbook": {}}
			with open(STATE_PATH, "r", encoding="utf-8") as f:
				data = json.load(f)
			if not isinstance(data, dict):
				return {"version": 1, "rollcalls": {}, "workbook": {}}
			data.setdefault("version", 1)
			data.setdefault("rollcalls", {})
			data.setdefault("workbook", {})
			return data
		except Exception:
			logger.warning("Failed to load rollcall state; starting fresh.", exc_info=True)
			return {"version": 1, "rollcalls": {}, "workbook": {}}

	def _save_state(self) -> None:
		try:
			self._state["updated_at"] = datetime.utcnow().isoformat()
			os.makedirs(os.path.dirname(STATE_PATH) or ".", exist_ok=True)
			tmp_path = f"{STATE_PATH}.tmp"
			with open(tmp_path, "w", encoding="utf-8") as f:
				json.dump(self._state, f, indent=2, ensure_ascii=False)
			os.replace(tmp_path, STATE_PATH)
		except Exception:
			logger.warning("Failed to save rollcall state.", exc_info=True)

	def _rc_state(self, key: str) -> dict:
		rc = self._state.setdefault("rollcalls", {}).setdefault(key, {})
		rc.setdefault("current_week", None)
		rc.setdefault("rollcall_message_id", None)
		rc.setdefault("rollcall_channel_id", None)
		rc.setdefault("html_message_id", None)
		rc.setdefault("html_channel_id", None)
		rc.setdefault("html_url", None)
		rc.setdefault("last_sent_at", None)
		return rc

	def _get_cfg_for_rollcall_message(self, message_id: int) -> tuple[Optional[RollCallConfig], Optional[dict]]:
		for cfg in ROLLCALLS:
			st = self._rc_state(cfg.key)
			if st.get("rollcall_message_id") == message_id:
				return cfg, st
		return None, None

	def _is_rollcall_locked(self, st: dict) -> bool:
		"""Return True if the roll call should be locked based on last_sent_at."""
		hours = ROLLCALL_LOCK_HOURS
		if hours is None:
			return False
		try:
			hours_f = float(hours)
		except Exception:
			return False
		if hours_f <= 0:
			return False

		sent_at = st.get("last_sent_at")
		if not isinstance(sent_at, str) or not sent_at.strip():
			return False
		try:
			dt = datetime.fromisoformat(sent_at.strip())
		except Exception:
			return False
		# We store UTC timestamps without tzinfo; treat as UTC.
		if dt.tzinfo is not None:
			dt = dt.astimezone(tz=None).replace(tzinfo=None)
		return (datetime.utcnow() - dt) >= timedelta(hours=hours_f)

	def _workbook_state(self) -> dict:
		wb = self._state.setdefault("workbook", {})
		wb.setdefault("message_id", None)
		wb.setdefault("channel_id", None)
		wb.setdefault("url", None)
		wb.setdefault("updated_at", None)
		return wb

	def _workbook_upload_channel_id(self) -> Optional[int]:
		if WORKBOOK_UPLOAD_CHANNEL_ID:
			return int(WORKBOOK_UPLOAD_CHANNEL_ID)
		if HTML_CHANNEL_ID:
			return int(HTML_CHANNEL_ID)
		if ROLLCALLS:
			return int(ROLLCALLS[0].channel_id)
		return None

	async def _post_workbook(self) -> Optional[str]:
		"""Upload the current rollcall workbook and return the attachment URL."""
		state = self._workbook_state()
		channel_id = self._workbook_upload_channel_id()
		if not channel_id:
			return state.get("url")
		channel = await self._get_text_channel(channel_id)
		if not channel:
			return state.get("url")
		if not os.path.exists(WORKBOOK_PATH):
			return state.get("url")

		old_id = state.get("message_id")
		old_channel_id = state.get("channel_id")
		if isinstance(old_id, int):
			try:
				old_channel = channel
				if isinstance(old_channel_id, int) and old_channel_id != channel.id:
					fetched_old = await self._get_text_channel(old_channel_id)
					if fetched_old:
						old_channel = fetched_old
				old_msg = await old_channel.fetch_message(old_id)
				await old_msg.delete()
			except discord.NotFound:
				pass
			except discord.Forbidden:
				logger.warning("RollCall: missing permission to delete old workbook message")
			except Exception:
				logger.warning("RollCall: failed deleting old workbook message", exc_info=True)

		try:
			file = discord.File(fp=WORKBOOK_PATH, filename="rollcall.xlsx")
			msg = await channel.send(content="Rollcall workbook (XLSX)", file=file)
			state["message_id"] = msg.id
			state["channel_id"] = channel.id
			state["url"] = msg.attachments[0].url if msg.attachments else None
			state["updated_at"] = datetime.utcnow().isoformat()
			return state["url"]
		except Exception:
			logger.warning("RollCall: failed uploading workbook", exc_info=True)
			return state.get("url")

	# -----------------
	# Scheduler
	# -----------------
	def _start_scheduler(self) -> None:
		try:
			import pytz

			tz = pytz.timezone(TIMEZONE)
		except Exception:
			logger.warning("Invalid TIMEZONE %r; falling back to UTC", TIMEZONE)
			import pytz

			tz = pytz.UTC

		# Avoid starting twice.
		if self._scheduler and getattr(self._scheduler, "running", False):
			return

		# Bind to the currently running loop (prevents jobs running without a loop).
		try:
			loop = asyncio.get_running_loop()
		except RuntimeError:
			# If we somehow got here without a running loop, bail; cog_load should call us with a loop.
			logger.error("RollCall scheduler start requested with no running event loop")
			return

		self._scheduler = AsyncIOScheduler(timezone=tz, event_loop=loop)

		trigger = CronTrigger(
			day_of_week=SCHEDULE_WEEKDAY,
			hour=SCHEDULE_HOUR,
			minute=SCHEDULE_MINUTE,
			timezone=tz,
		)

		# Schedule the coroutine directly on the asyncio scheduler.
		self._scheduler.add_job(self._scheduled_send, trigger=trigger, id="weekly_rollcall", replace_existing=True)
		self._scheduler.start()

	async def _backstop_loop(self) -> None:
		await self.bot.wait_until_ready()
		while True:
			try:
				await asyncio.sleep(float(BACKSTOP_REFRESH_MINUTES) * 60.0)
				await self._refresh_all(reason="backstop")
			except asyncio.CancelledError:
				return
			except Exception:
				logger.exception("RollCall backstop refresh failed")

	def _debounced_refresh(self, *, reason: str, delay_seconds: float = 3.0) -> None:
		if self._debounce_task and not self._debounce_task.done():
			self._debounce_task.cancel()
		self._debounce_task = asyncio.create_task(self._debounce_worker(reason=reason, delay_seconds=delay_seconds))

	async def _debounce_worker(self, *, reason: str, delay_seconds: float) -> None:
		try:
			await asyncio.sleep(delay_seconds)
			await self._refresh_all(reason=reason)
		except asyncio.CancelledError:
			return

	# -----------------
	# Helpers
	# -----------------
	def _tracked_role_ids(self, cfg: RollCallConfig) -> list[int]:
		ids: list[int] = []
		tri = cfg.tracked_role_ids
		if tri is not None:
			# Allow configs to accidentally pass a single int, and normalize it.
			if isinstance(tri, int):
				candidates = [tri]
			elif isinstance(tri, str):
				candidates = []
			else:
				try:
					candidates = list(tri)
				except TypeError:
					candidates = []
			for rid in candidates:
				try:
					ids.append(int(rid))
				except Exception:
					continue
		if cfg.tracked_role_id:
			try:
				ids.append(int(cfg.tracked_role_id))
			except Exception:
				pass
		# de-dupe while keeping order
		seen: set[int] = set()
		out: list[int] = []
		for rid in ids:
			if rid not in seen:
				seen.add(rid)
				out.append(rid)
		return out

	def _ping_role_ids(self, cfg: RollCallConfig) -> list[int]:
		ids: list[int] = []
		pri = cfg.ping_role_ids
		if pri is not None:
			if isinstance(pri, int):
				candidates = [pri]
			elif isinstance(pri, str):
				candidates = []
			else:
				try:
					candidates = list(pri)
				except TypeError:
					candidates = []
			for rid in candidates:
				try:
					ids.append(int(rid))
				except Exception:
					continue
		if cfg.ping_role_id:
			try:
				ids.append(int(cfg.ping_role_id))
			except Exception:
				pass
		seen: set[int] = set()
		out: list[int] = []
		for rid in ids:
			if rid not in seen:
				seen.add(rid)
				out.append(rid)
		return out

	def _ping_mentions(self, cfg: RollCallConfig) -> str:
		ids = self._ping_role_ids(cfg)
		if not ids:
			return ""
		return "".join(f"<@&{rid}> " for rid in ids)

	def _apply_partial_tick_markers(self, guild: discord.Guild, wb: Workbook, *, week_label: str) -> None:
		"""If a member ticks at least one roll call this week, mark 🅾️ on other roll calls they missed."""
		PARTIAL = "🅾️"

		active_cfgs: list[RollCallConfig] = []
		for cfg in ROLLCALLS:
			st = self._rc_state(cfg.key)
			if st.get("current_week") == week_label:
				active_cfgs.append(cfg)

		# Only meaningful when multiple roll calls are active for the same week.
		if len(active_cfgs) < 2:
			return

		# Who has ticked *any* roll call this week?
		ticked_anywhere: set[int] = set()
		for cfg in active_cfgs:
			ws = self._get_or_create_sheet(wb, cfg)
			status = self._get_week_status(ws, week_label)
			for uid, v in status.items():
				if v == "✅":
					ticked_anywhere.add(uid)

		for cfg in active_cfgs:
			ws = self._get_or_create_sheet(wb, cfg)
			headers = self._sheet_headers(ws)
			if week_label not in headers:
				continue
			week_col = headers.index(week_label) + 1

			expected = self._expected_members(guild, cfg)
			if not expected:
				continue

			for m in expected:
				row = self._upsert_member_row(ws, m.id, m.display_name)
				cur = ws.cell(row=row, column=week_col).value
				cur_s = str(cur) if cur is not None else ""

				# Only replace explicit misses (❌) with partial marker.
				if cur_s == "❌" and m.id in ticked_anywhere:
					ws.cell(row=row, column=week_col, value=PARTIAL)
				elif cur_s == PARTIAL and m.id not in ticked_anywhere:
					ws.cell(row=row, column=week_col, value="❌")
	async def _get_text_channel(self, channel_id: int) -> Optional[discord.TextChannel]:
		ch = self.bot.get_channel(channel_id)
		if isinstance(ch, discord.TextChannel):
			return ch
		try:
			fetched = await self.bot.fetch_channel(channel_id)
			return fetched if isinstance(fetched, discord.TextChannel) else None
		except Exception:
			return None

	def _week_label(self, d: date) -> str:
		# Columns are week numbers with the date rollcall was sent.
		week = d.isocalendar().week
		return f"W{week:02d} {d.strftime('%d/%m/%Y')}"

	def _migrate_week_headers_to_ddmmyyyy(self, ws) -> None:
		"""Convert legacy 'Wxx YYYY-MM-DD' headers to 'Wxx DD/MM/YYYY' in-place."""
		try:
			for col in range(3, ws.max_column + 1):
				cell = ws.cell(row=1, column=col)
				val = cell.value
				if not isinstance(val, str):
					continue
				m = re.match(r"^W(\d{2})\s+(\d{4})-(\d{2})-(\d{2})$", val.strip())
				if not m:
					continue
				week = m.group(1)
				yy = m.group(2)
				mm = m.group(3)
				dd = m.group(4)
				cell.value = f"W{week} {dd}/{mm}/{yy}"
		except Exception:
			logger.warning("Failed migrating week headers to DD/MM/YYYY", exc_info=True)

	def _rollcall_date_for_now(self) -> date:
		"""Return the rollcall date corresponding to the most recent scheduled send."""
		try:
			import pytz

			tz = pytz.timezone(TIMEZONE)
		except Exception:
			import pytz

			tz = pytz.UTC

		now_local = datetime.now(tz)
		scheduled_t = time(hour=SCHEDULE_HOUR, minute=SCHEDULE_MINUTE)
		# Find most recent weekday occurrence
		weekday_map = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
		target_wd = weekday_map.get(str(SCHEDULE_WEEKDAY).lower(), 0)
		days_back = (now_local.weekday() - target_wd) % 7
		candidate = (now_local.date() - timedelta(days=days_back))
		if days_back == 0 and now_local.time() < scheduled_t:
			candidate = candidate - timedelta(days=7)
		return candidate

	def _maybe_import_legacy_xls(self) -> None:
		if not IMPORT_LEGACY_XLS_PATH:
			return
		if os.path.exists(WORKBOOK_PATH):
			return
		if not os.path.exists(IMPORT_LEGACY_XLS_PATH):
			logger.warning("Legacy .xls import path not found: %s", IMPORT_LEGACY_XLS_PATH)
			return

		try:
			import pandas as pd
		except Exception:
			logger.warning("pandas not installed; cannot import legacy .xls")
			return

		try:
			xls = pd.ExcelFile(IMPORT_LEGACY_XLS_PATH)
			wb = Workbook()
			# Remove default sheet
			default_sheet = wb.active
			wb.remove(default_sheet)

			for sheet_name in xls.sheet_names:
				df = xls.parse(sheet_name)
				ws = wb.create_sheet(title=str(sheet_name)[:31])
				# Write headers
				ws.append([str(c) for c in df.columns.tolist()])
				# Write rows
				for row in df.itertuples(index=False):
					ws.append([None if (isinstance(v, float) and pd.isna(v)) else v for v in row])

			os.makedirs(os.path.dirname(WORKBOOK_PATH) or ".", exist_ok=True)
			wb.save(WORKBOOK_PATH)
			logger.info("Imported legacy .xls into %s", WORKBOOK_PATH)
		except Exception:
			logger.exception("Failed importing legacy .xls")

	# -----------------
	# Workbook (.xlsx)
	# -----------------
	def _load_or_create_workbook(self) -> Workbook:
		self._maybe_import_legacy_xls()

		if os.path.exists(WORKBOOK_PATH):
			return load_workbook(WORKBOOK_PATH)

		wb = Workbook()
		# Keep the default sheet but rename it to something sane.
		wb.active.title = "README"
		wb.active.append(["This workbook is managed by the bot."])
		os.makedirs(os.path.dirname(WORKBOOK_PATH) or ".", exist_ok=True)
		wb.save(WORKBOOK_PATH)
		return wb

	def _get_or_create_sheet(self, wb: Workbook, cfg: RollCallConfig):
		name = cfg.key[:31]
		if name in wb.sheetnames:
			ws = wb[name]
		else:
			ws = wb.create_sheet(title=name)
			ws.append(["User ID", "Nickname"])
		self._migrate_week_headers_to_ddmmyyyy(ws)
		return ws

	def _sheet_headers(self, ws) -> list[str]:
		headers: list[str] = []
		for cell in ws[1]:
			v = cell.value
			if v is None:
				headers.append("")
			elif isinstance(v, str):
				headers.append(v.strip())
			else:
				headers.append(str(v))
		return headers

	def _parse_week_header(self, header: str) -> Optional[tuple[int, date]]:
		"""Parse 'Wxx DD/MM/YYYY' or legacy 'Wxx YYYY-MM-DD' into (week, date)."""
		if not header:
			return None
		h = header.strip()
		m = re.match(r"^W(\d{2})\s+(\d{2})/(\d{2})/(\d{4})$", h)
		if m:
			week = int(m.group(1))
			dd = int(m.group(2))
			mm = int(m.group(3))
			yy = int(m.group(4))
			return (week, date(yy, mm, dd))
		m = re.match(r"^W(\d{2})\s+(\d{4})-(\d{2})-(\d{2})$", h)
		if m:
			week = int(m.group(1))
			yy = int(m.group(2))
			mm = int(m.group(3))
			dd = int(m.group(4))
			return (week, date(yy, mm, dd))
		return None

	def _ensure_week_column(self, ws, week_label: str) -> int:
		headers = self._sheet_headers(ws)
		canonical = week_label.strip() if isinstance(week_label, str) else str(week_label)
		if canonical in headers:
			return headers.index(canonical) + 1

		# Match by normalized (week, date) to avoid duplicate columns when formats differ.
		wanted = self._parse_week_header(canonical)
		if wanted:
			for idx, h in enumerate(headers):
				parsed = self._parse_week_header(h)
				if parsed and parsed == wanted:
					col = idx + 1
					# Rename header to canonical to prevent future duplicates
					ws.cell(row=1, column=col, value=canonical)
					return col

		# No match: create new column
		ws.cell(row=1, column=len(headers) + 1, value=canonical)
		return len(headers) + 1

	def _index_user_rows(self, ws) -> dict[str, int]:
		idx: dict[str, int] = {}
		for r in range(2, ws.max_row + 1):
			v = ws.cell(row=r, column=1).value
			if v is None:
				continue
			idx[str(v)] = r
		return idx

	def _upsert_member_row(self, ws, user_id: int, nickname: str) -> int:
		idx = self._index_user_rows(ws)
		key = str(user_id)
		if key in idx:
			row = idx[key]
			ws.cell(row=row, column=2, value=nickname)
			return row
		row = ws.max_row + 1
		ws.cell(row=row, column=1, value=key)
		ws.cell(row=row, column=2, value=nickname)
		return row

	def _set_cell(self, ws, row: int, col: int, value: str) -> None:
		ws.cell(row=row, column=col, value=value)

	def _save_workbook(self, wb: Workbook) -> None:
		tmp_path = f"{WORKBOOK_PATH}.tmp"
		wb.save(tmp_path)
		os.replace(tmp_path, WORKBOOK_PATH)

	# -----------------
	# Core flows
	# -----------------
	async def _scheduled_send(self) -> None:
		await self.bot.wait_until_ready()
		await self._send_rollcalls(reason="scheduled")

	async def _send_rollcalls(self, *, reason: str) -> None:
		async with self._lock:
			guild = self.bot.get_guild(GUILD_ID)
			if not guild:
				logger.warning("RollCall: guild not found")
				return

			rollcall_d = self._rollcall_date_for_now()
			week_label = self._week_label(rollcall_d)

			wb = self._load_or_create_workbook()

			for cfg in ROLLCALLS:
				try:
					await self._send_rollcall_for_cfg(guild, wb, cfg, week_label=week_label, rollcall_d=rollcall_d, reason=reason)
				except Exception:
					logger.exception("RollCall: failed sending for %s", cfg.key)

			self._save_workbook(wb)
			# Upload workbook for download link, then refresh embeds to include latest URL.
			workbook_url = await self._post_workbook()
			if workbook_url:
				for cfg in ROLLCALLS:
					try:
						await self._update_outputs_for_cfg(
							guild,
							wb,
							cfg,
							week_label=week_label,
							reason="workbook_link",
							update_html=False,
						)
					except Exception:
						logger.exception("RollCall: failed workbook embed refresh for %s", cfg.key)
			self._save_state()

	async def _refresh_all(self, *, reason: str) -> None:
		async with self._lock:
			guild = self.bot.get_guild(GUILD_ID)
			if not guild:
				return
			week_label = self._week_label(self._rollcall_date_for_now())
			wb = self._load_or_create_workbook()
			self._apply_partial_tick_markers(guild, wb, week_label=week_label)
			prev_workbook_url = self._workbook_state().get("url")
			update_html = (reason != "reaction") or UPDATE_HTML_ON_REACTION
			for cfg in ROLLCALLS:
				try:
					await self._update_outputs_for_cfg(
						guild,
						wb,
						cfg,
						week_label=week_label,
						reason=reason,
						update_html=update_html,
					)
				except Exception:
					logger.exception("RollCall: failed refresh for %s", cfg.key)
			self._save_workbook(wb)

			should_upload_workbook = True
			if reason == "reaction" and not UPLOAD_WORKBOOK_ON_REACTION:
				should_upload_workbook = False
			if should_upload_workbook:
				new_workbook_url = await self._post_workbook()
				if new_workbook_url and new_workbook_url != prev_workbook_url:
					for cfg in ROLLCALLS:
						try:
							await self._update_outputs_for_cfg(
								guild,
								wb,
								cfg,
								week_label=week_label,
								reason="workbook_link",
								update_html=False,
							)
						except Exception:
							logger.exception("RollCall: failed workbook embed refresh for %s", cfg.key)
			self._save_state()

	async def _send_rollcall_for_cfg(
		self,
		guild: discord.Guild,
		wb: Workbook,
		cfg: RollCallConfig,
		*,
		week_label: str,
		rollcall_d: date,
		reason: str,
	) -> None:
		channel = await self._get_text_channel(cfg.channel_id)
		if not channel:
			logger.warning("RollCall %s: channel not found", cfg.key)
			return

		state = self._rc_state(cfg.key)
		# If we already sent this week, don't re-post; just refresh embed/html.
		if state.get("current_week") == week_label and isinstance(state.get("rollcall_message_id"), int):
			# If the message was deleted manually, clear state so we can re-post.
			try:
				await channel.fetch_message(int(state["rollcall_message_id"]))
			except discord.NotFound:
				state["rollcall_message_id"] = None
				state["rollcall_channel_id"] = None
				state["html_message_id"] = None
				state["html_channel_id"] = None
				state["html_url"] = None
			else:
				await self._update_outputs_for_cfg(guild, wb, cfg, week_label=week_label, reason="already_sent", update_html=True)
				return

		ws = self._get_or_create_sheet(wb, cfg)
		week_col = self._ensure_week_column(ws, week_label)

		expected_members = self._expected_members(guild, cfg)
		for m in expected_members:
			row = self._upsert_member_row(ws, m.id, m.display_name)
			self._set_cell(ws, row, week_col, "❌")

		# Keep excluded members visible in the table (bottom section), but don't mark them ❌.
		for m in self._excluded_members(guild):
			self._upsert_member_row(ws, m.id, m.display_name)

		ping = self._ping_mentions(cfg)
		workbook_url = self._workbook_state().get("url")
		embed = await self._build_status_embed(
			guild,
			ws,
			cfg,
			week_label=week_label,
			rollcall_d=rollcall_d,
			html_url=state.get("html_url"),
			workbook_url=workbook_url,
		)

		msg = await channel.send(content=f"{ping}Weekly roll call: react with {ROLLCALL_EMOJI}", embed=embed)
		try:
			await msg.add_reaction(ROLLCALL_EMOJI)
		except Exception:
			logger.warning("RollCall %s: failed to add reaction", cfg.key, exc_info=True)

		state["current_week"] = week_label
		state["rollcall_message_id"] = msg.id
		state["rollcall_channel_id"] = channel.id
		state["last_sent_at"] = datetime.utcnow().isoformat()

		# Post HTML + refresh embed to include latest HTML link
		await self._update_outputs_for_cfg(guild, wb, cfg, week_label=week_label, reason=reason, update_html=True)

	def _expected_members(self, guild: discord.Guild, cfg: RollCallConfig) -> list[discord.Member]:
		role_ids = self._tracked_role_ids(cfg)
		if role_ids:
			members: dict[int, discord.Member] = {}
			for rid in role_ids:
				role = guild.get_role(rid)
				if not role:
					continue
				for m in role.members:
					if self._is_member_excluded(m):
						continue
					members[m.id] = m
			return sorted(members.values(), key=lambda m: (m.display_name or "").lower())
		# fallback: nobody "expected" (we'll still record reactions)
		return []

	async def _update_outputs_for_cfg(
		self,
		guild: discord.Guild,
		wb: Workbook,
		cfg: RollCallConfig,
		*,
		week_label: str,
		reason: str,
		update_html: bool,
	) -> None:
		state = self._rc_state(cfg.key)
		channel = await self._get_text_channel(cfg.channel_id)
		if not channel:
			return

		ws = self._get_or_create_sheet(wb, cfg)
		rollcall_d = self._parse_week_label_date(week_label) or self._rollcall_date_for_now()

		# Ensure member nicknames stay up to date in the sheet.
		for m in self._expected_members(guild, cfg):
			self._upsert_member_row(ws, m.id, m.display_name)
		for m in self._excluded_members(guild):
			self._upsert_member_row(ws, m.id, m.display_name)

		if update_html:
			await self._post_html_for_cfg(guild, ws, cfg, week_label=week_label)
		html_url = self._rc_state(cfg.key).get("html_url")

		workbook_url = self._workbook_state().get("url")
		embed = await self._build_status_embed(
			guild,
			ws,
			cfg,
			week_label=week_label,
			rollcall_d=rollcall_d,
			html_url=html_url,
			workbook_url=workbook_url,
		)
		msg_id = state.get("rollcall_message_id")
		if isinstance(msg_id, int):
			try:
				msg = await channel.fetch_message(msg_id)
				await msg.edit(embed=embed)
			except discord.NotFound:
				pass
			except discord.Forbidden:
				logger.warning("RollCall %s: missing permission to edit message", cfg.key)
			except Exception:
				logger.warning("RollCall %s: failed editing message", cfg.key, exc_info=True)

	def _parse_week_label_date(self, week_label: str) -> Optional[date]:
		# formats:
		# - 'W07 16/02/2026'
		# - legacy 'W07 2026-02-16'
		try:
			parts = str(week_label).split()
			if len(parts) >= 2:
				v = parts[1]
				try:
					return datetime.strptime(v, "%d/%m/%Y").date()
				except Exception:
					return date.fromisoformat(v)
			return None
		except Exception:
			return None

	async def _build_status_embed(
		self,
		guild: discord.Guild,
		ws,
		cfg: RollCallConfig,
		*,
		week_label: str,
		rollcall_d: date,
		html_url: Optional[str],
		workbook_url: Optional[str],
	) -> discord.Embed:
		expected = self._expected_members(guild, cfg)
		expected_ids = {m.id for m in expected}
		status = self._get_week_status(ws, week_label)

		ticked_ids = {uid for uid, v in status.items() if v == "✅"}
		missing_ids = (expected_ids - ticked_ids) if expected_ids else set()

		def name_for(uid: int) -> str:
			m = guild.get_member(uid)
			return m.display_name if m else str(uid)

		ticked_names = [name_for(uid) for uid in sorted(ticked_ids, key=lambda u: name_for(u).lower())]
		missing_names = [name_for(uid) for uid in sorted(missing_ids, key=lambda u: name_for(u).lower())]

		embed = discord.Embed(
			title=cfg.title,
			color=discord.Color.green(),
			timestamp=datetime.utcnow(),
			description=(
				f"**Week:** {week_label}\n"
				f"**Roll call date:** {rollcall_d.strftime('%d/%m/%Y')}\n"
				f"**React with:** {ROLLCALL_EMOJI}"
			),
			url=(html_url or None),
		)

		if html_url:
			embed.description += f"\n\n[Open full table (HTML)]({html_url})"
		if workbook_url:
			embed.description += "\nRollcall workbook (XLSX) is posted in the admin channel — ask an admin if you need it."

		if cfg.embed_image_url:
			url = str(cfg.embed_image_url).strip()
			if url.startswith("http://") or url.startswith("https://"):
				embed.set_image(url=url)

		embed.add_field(name="Ticked", value=self._chunk_list(ticked_names), inline=False)
		if expected_ids:
			embed.add_field(name="Missing", value=self._chunk_list(missing_names) if missing_names else "None", inline=False)
		else:
			embed.add_field(name="Missing", value="(No tracked role configured)", inline=False)

		embed.set_footer(text="Updates live as reactions are added/removed")
		return embed

	def _chunk_list(self, items: list[str], *, max_chars: int = 1024) -> str:
		"""Join items into a string guaranteed to be <= max_chars (Discord embed field limit)."""
		if not items:
			return "None"

		out = ""
		for i, s in enumerate(items):
			sep = ", " if out else ""
			candidate = out + sep + s
			if len(candidate) <= max_chars:
				out = candidate
				continue

			# Doesn't fit: append an ellipsis and the number of omitted items (when possible).
			remaining = len(items) - i
			suffix = (", … (+%d more)" % remaining) if out else ("… (+%d more)" % remaining)
			if out:
				if len(out) + len(suffix) <= max_chars:
					out = out + suffix
				else:
					short_suffix = ", …"
					if len(out) + len(short_suffix) <= max_chars:
						out = out + short_suffix
					else:
						out = (out[: max(0, max_chars - 1)] + "…")[:max_chars]
			else:
				# First item alone doesn't fit (very long nickname) -> truncate.
				out = (str(s)[: max(0, max_chars - 1)] + "…")[:max_chars]
			break

		return out

	def _get_week_status(self, ws, week_label: str) -> dict[int, str]:
		headers = self._sheet_headers(ws)
		if week_label not in headers:
			return {}
		col = headers.index(week_label) + 1
		out: dict[int, str] = {}
		for r in range(2, ws.max_row + 1):
			uid = ws.cell(row=r, column=1).value
			if uid is None:
				continue
			v = ws.cell(row=r, column=col).value
			try:
				out[int(str(uid))] = str(v) if v is not None else ""
			except Exception:
				continue
		return out

	async def _post_html_for_cfg(self, guild: discord.Guild, ws, cfg: RollCallConfig, *, week_label: str) -> None:
		state = self._rc_state(cfg.key)
		target_channel_id = HTML_CHANNEL_ID if HTML_CHANNEL_ID else cfg.channel_id
		channel = await self._get_text_channel(int(target_channel_id))
		if not channel:
			return

		# Delete previous HTML message to keep clean
		old_id = state.get("html_message_id")
		old_channel_id = state.get("html_channel_id")
		if isinstance(old_id, int):
			try:
				old_channel = channel
				if isinstance(old_channel_id, int) and old_channel_id != channel.id:
					fetched_old = await self._get_text_channel(old_channel_id)
					if fetched_old:
						old_channel = fetched_old
				old_msg = await old_channel.fetch_message(old_id)
				await old_msg.delete()
			except discord.NotFound:
				pass
			except discord.Forbidden:
				logger.warning("RollCall %s: missing permission to delete old HTML message", cfg.key)
			except Exception:
				logger.warning("RollCall %s: failed deleting old HTML message", cfg.key, exc_info=True)

		html_text = self._render_html(guild, ws, cfg, highlight_week=week_label)
		file_bytes = html_text.encode("utf-8")
		file = discord.File(fp=io.BytesIO(file_bytes), filename=f"{cfg.key}_rollcall.html")

		msg = await channel.send(content=f"{cfg.title} (HTML table)", file=file)
		state["html_message_id"] = msg.id
		state["html_channel_id"] = channel.id
		state["html_url"] = msg.attachments[0].url if msg.attachments else None

	def _render_html(self, guild: discord.Guild, ws, cfg: RollCallConfig, *, highlight_week: Optional[str]) -> str:
		headers = self._sheet_headers(ws)
		if not headers or headers[0] != "User ID":
			headers = ["User ID", "Nickname"] + headers[2:]

		# Display headers: hide User ID, add Flags column after Nickname
		week_headers = headers[2:]
		head_cells = ["Nickname", "Flags"] + [str(h) for h in week_headers]
		head_html = "".join(f"<th>{html.escape(h)}</th>" for h in head_cells)

		def parse_uid(v) -> Optional[int]:
			try:
				if v is None:
					return None
				return int(str(v).strip())
			except Exception:
				return None

		def row_flags(uid_int: Optional[int]) -> list[str]:
			if uid_int is None:
				return ["LEFT"]
			m = guild.get_member(uid_int)
			if not m:
				return ["LEFT"]
			return self._member_exclusion_markers(m)

		main_rows: list[str] = []
		excluded_rows: list[str] = []

		for r in range(2, ws.max_row + 1):
			uid_val = ws.cell(row=r, column=1).value
			nick_val = ws.cell(row=r, column=2).value
			if uid_val is None and nick_val is None:
				continue
			uid_int = parse_uid(uid_val)
			flags = row_flags(uid_int)
			flags_str = ", ".join(flags) if flags else ""

			# Use stored nickname (keeps history) but it should be kept up-to-date by refresh.
			nick = str(nick_val or "")
			cells = [html.escape(nick), html.escape(flags_str)]
			for c in range(3, len(headers) + 1):
				v = ws.cell(row=r, column=c).value
				cells.append(html.escape(str(v or "")))
			row_html = "<tr>" + "".join(f"<td>{v}</td>" for v in cells) + "</tr>"

			# Decide which table
			is_excluded = False
			if "LEFT" in flags:
				is_excluded = True
			if any(f in ("HG", "AWOL") for f in flags):
				is_excluded = True

			if is_excluded:
				excluded_rows.append(row_html)
			else:
				main_rows.append(row_html)

		main_table = "".join(main_rows) if main_rows else '<tr><td colspan="100">No active members.</td></tr>'
		excluded_table = "".join(excluded_rows) if excluded_rows else '<tr><td colspan="100">None.</td></tr>'

		highlight_css = ""
		if highlight_week and highlight_week in headers:
			idx = headers.index(highlight_week)  # 0-based in headers
			# Display columns are: Nickname(1), Flags(2), then week columns starting at headers[2]
			display_col = (idx - 2) + 3
			# nth-child is 1-based; td/th
			nth = display_col + 1
			highlight_css = f"th:nth-child({nth}), td:nth-child({nth}) {{ background: #fff7cc; }}"

		return f"""<!doctype html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width,initial-scale=1\" />
  <title>{html.escape(cfg.title)}</title>
  <style>
	body {{ font-family: Arial, sans-serif; padding: 16px; }}
	h1 {{ margin: 0 0 12px 0; }}
	table {{ border-collapse: collapse; width: 100%; }}
	th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
	th {{ background: #f5f5f5; position: sticky; top: 0; }}
	tr:nth-child(even) {{ background: #fafafa; }}
	{highlight_css}
  </style>
</head>
<body>
  <h1>{html.escape(cfg.title)}</h1>
	  <p><strong>Flags:</strong> HG = Homeguard, AWOL = AWOL, LEFT = left server</p>
  <p>Last updated: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}</p>
	  <h2>Roll call</h2>
	  <table>
	    <thead><tr>{head_html}</tr></thead>
	    <tbody>{main_table}</tbody>
	  </table>

	  <h2 style="margin-top: 24px;">Excluded / Inactive</h2>
	  <table>
	    <thead><tr>{head_html}</tr></thead>
	    <tbody>{excluded_table}</tbody>
	  </table>
</body>
</html>"""

	# -----------------
	# Events
	# -----------------
	@commands.Cog.listener()
	async def on_ready(self):
		# One-time refresh after login.
		if self._refresh_task is None or self._refresh_task.done():
			self._refresh_task = asyncio.create_task(self._refresh_all(reason="startup"))

	@commands.Cog.listener()
	async def on_raw_reaction_add(self, payload: discord.RawReactionActionEvent):
		if payload.guild_id != GUILD_ID:
			return
		if payload.user_id == getattr(self.bot.user, "id", None):
			return

		# Only enforce reactions on *our* roll call post(s).
		match_cfg, st = self._get_cfg_for_rollcall_message(payload.message_id)
		if not match_cfg or not st:
			return

		# Allow only the configured tick emoji.
		if str(payload.emoji) == str(ROLLCALL_EMOJI):
			# Enforce rollcall lock.
			if self._is_rollcall_locked(st):
				guild = self.bot.get_guild(GUILD_ID)
				member = guild.get_member(payload.user_id) if guild else None
				if not member and guild:
					try:
						member = await guild.fetch_member(payload.user_id)
					except Exception:
						member = None

				# Best-effort: remove the reaction they added.
				try:
					ch = await self._get_text_channel(payload.channel_id) if payload.channel_id else None
					if ch and isinstance(member, discord.Member):
						msg = await ch.fetch_message(payload.message_id)
						await msg.remove_reaction(payload.emoji, member)
				except Exception:
					pass

				# DM user
				if isinstance(member, discord.Member):
					try:
						await member.send(
							"this roll call is locked sorry :( make be sure to tick next week!"
						)
					except Exception:
						pass
				return

			await self._handle_reaction_change(payload, marked=True)
			return

		guild = self.bot.get_guild(GUILD_ID)
		member = guild.get_member(payload.user_id) if guild else None
		if not member and guild:
			try:
				member = await guild.fetch_member(payload.user_id)
			except Exception:
				member = None

		# Best-effort: remove the invalid reaction.
		try:
			ch = await self._get_text_channel(payload.channel_id) if payload.channel_id else None
			if ch and isinstance(member, discord.Member):
				msg = await ch.fetch_message(payload.message_id)
				await msg.remove_reaction(payload.emoji, member)
		except Exception:
			pass

		# DM user
		if isinstance(member, discord.Member):
			try:
				await member.send("you must tick only :( reaction removed")
			except Exception:
				pass

	@commands.Cog.listener()
	async def on_raw_reaction_remove(self, payload: discord.RawReactionActionEvent):
		if payload.guild_id != GUILD_ID:
			return
		if str(payload.emoji) != str(ROLLCALL_EMOJI):
			return

		# If the rollcall is locked, ignore removes so attendance can't be changed after the deadline.
		_cfg, st = self._get_cfg_for_rollcall_message(payload.message_id)
		if st and self._is_rollcall_locked(st):
			return
		await self._handle_reaction_change(payload, marked=False)

	async def _handle_reaction_change(self, payload: discord.RawReactionActionEvent, *, marked: bool) -> None:
		# Find which rollcall this message belongs to
		match_cfg: Optional[RollCallConfig] = None
		week_label: Optional[str] = None
		for cfg in ROLLCALLS:
			st = self._rc_state(cfg.key)
			if st.get("rollcall_message_id") == payload.message_id:
				match_cfg = cfg
				week_label = st.get("current_week")
				break
		if not match_cfg or not week_label:
			return

		guild = self.bot.get_guild(GUILD_ID)
		if not guild:
			return

		member = guild.get_member(payload.user_id)
		if not member:
			try:
				member = await guild.fetch_member(payload.user_id)
			except Exception:
				member = None

		# Enforce role gating.
		# If tracked roles are set, only members with any of those roles can tick/un-tick this roll call.
		# This MUST apply to reaction removes too, because when we remove an invalid reaction we will also
		# receive the corresponding raw_reaction_remove event.
		required_role_ids = set(self._tracked_role_ids(match_cfg))
		if required_role_ids:
			if not isinstance(member, discord.Member):
				return
			if not any(r.id in required_role_ids for r in member.roles):
				if marked:
					# Best-effort: remove the reaction they added.
					try:
						ch = await self._get_text_channel(payload.channel_id) if payload.channel_id else None
						if ch:
							msg = await ch.fetch_message(payload.message_id)
							await msg.remove_reaction(payload.emoji, member)
					except Exception:
						# Not fatal (requires Manage Messages to remove others' reactions)
						pass

					# DM user
					try:
						await member.send(
							"You have reacted to the wrong roll call and don't have that role :( please contact an admin"
						)
					except Exception:
						pass
				return

		async with self._lock:
			wb = self._load_or_create_workbook()
			ws = self._get_or_create_sheet(wb, match_cfg)
			col = self._ensure_week_column(ws, week_label)
			row = self._upsert_member_row(ws, payload.user_id, member.display_name if member else str(payload.user_id))
			self._set_cell(ws, row, col, "✅" if marked else "❌")
			self._save_workbook(wb)
			self._save_state()

		# Outside lock: update outputs (debounced).
		self._debounced_refresh(reason="reaction", delay_seconds=REACTION_REFRESH_DEBOUNCE_SECONDS)

	@commands.Cog.listener()
	async def on_member_update(self, before: discord.Member, after: discord.Member):
		if before.guild.id != GUILD_ID:
			return
		if before.display_name != after.display_name:
			# Keep nicknames in workbook + HTML in sync.
			self._debounced_refresh(reason="nickname", delay_seconds=5.0)


async def setup(bot: commands.Bot):
	await bot.add_cog(RollCallCog(bot))
	logger.info("RollCallCog loaded")

