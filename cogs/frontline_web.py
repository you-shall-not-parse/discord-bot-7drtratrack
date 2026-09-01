from __future__ import annotations

import asyncio
import hmac
import html
import io
import json
import logging
import os
import re
import secrets
import sqlite3
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote

from aiohttp import ClientSession, ClientTimeout, web
from openpyxl import Workbook

from config import MAIN_GUILD_ID
from rank_order import DEFAULT_RANK_ORDER
from spreadsheet_security import safe_spreadsheet_value

logger = logging.getLogger(__name__)

FRONTEND_DIR = Path(__file__).resolve().parent.parent / "liberationapp" / "frontend"
CLAN_EMBLEM_PATH = Path(__file__).resolve().parent.parent / "data" / "emblem_7dr.png"
WEB_HOST = os.getenv("FRONTLINE_WEB_HOST", "127.0.0.1")
WEB_PORT = int(os.getenv("FRONTLINE_WEB_PORT", "7020"))
SESSION_COOKIE = "hll_frontline_session"
SESSION_SECONDS = 24 * 60 * 60
LOGIN_WINDOW_SECONDS = 15 * 60
LOGIN_MAX_FAILURES = 5
TURNSTILE_VERIFY_URL = "https://challenges.cloudflare.com/turnstile/v0/siteverify"
TURNSTILE_SITE_KEY = "0x4AAAAAAEjLElnYaILQaVYk"
TURNSTILE_HOSTNAME = "hllfrontline.com"
LOGIN_NAME_MAX_LENGTH = 80
HLLV_SEARCH_MAX_LENGTH = 64
HLLV_SEARCH_MIN_LENGTH = 2
HLLV_SEARCH_WINDOW_SECONDS = 60
HLLV_SEARCH_MAX_REQUESTS = 30
RAT_OF_THE_WEEK_ROLE_ID = 1461087295930106020
SERVER_STATUS_CACHE_SECONDS = 45
EXTERNAL_LINKS = {
    "bifrost": "https://frostbite.bifrostgaming.com/hll/guilds/7DR",
    "history": "https://7drhistostats.hllfrontline.com/",
    "merch": "https://7dr-hll-merch.myshopify.com/",
}


@dataclass(frozen=True)
class WebSession:
    expires_at: int
    claimed_name: str


class FrontlineWeb:
    """Small aiohttp site that presents live data owned by the Discord cogs."""

    def __init__(self, bot):
        self.bot = bot
        self._runner: web.AppRunner | None = None
        self._site: web.TCPSite | None = None
        self._app_pin = os.getenv("APPPIN", "")
        self._turnstile_site_key = TURNSTILE_SITE_KEY
        self._turnstile_secret_key = (
            os.getenv("TURNSTILE_SECRET", "").strip()
            or os.getenv("TURNSTILE_SECRET_KEY", "").strip()
        )
        self._turnstile_hostname = TURNSTILE_HOSTNAME
        self._sessions: dict[str, WebSession] = {}
        self._login_failures: dict[str, list[float]] = {}
        self._hllv_searches: dict[str, list[float]] = {}
        self._server_status_cache: tuple[float, list[dict[str, Any]]] = (0.0, [])
        self._server_status_lock = asyncio.Lock()

    async def start(self) -> None:
        if len(self._app_pin) < 8:
            raise RuntimeError("APPPIN must be set to at least 8 characters in the environment or .env file")
        if not self._turnstile_secret_key:
            raise RuntimeError("TURNSTILE_SECRET must be set in the environment or .env file")

        app = web.Application(middlewares=[self._security_headers, self._pin_auth], client_max_size=64 * 1024)
        app.router.add_get("/api/health", self.health)
        app.router.add_get("/login", self.login_page)
        app.router.add_post("/login", self.login)
        app.router.add_post("/logout", self.logout)
        app.router.add_get("/api/dashboard", self.dashboard)
        app.router.add_get("/api/hllv-search", self.hllv_search)
        app.router.add_get("/assets/{filename}", self.asset)
        app.router.add_get("/exports/rollcalls/{key}.html", self.rollcall_html_export)
        app.router.add_get("/exports/rollcalls/{key}.xlsx", self.rollcall_excel_export)
        app.router.add_get("/exports/trainees/{key}.html", self.trainee_html_export)
        app.router.add_get("/exports/trainees/{key}.xlsx", self.trainee_excel_export)
        app.router.add_get("/rollcalls/{key}", self.report_page)
        app.router.add_get("/trainees/{key}", self.report_page)
        app.router.add_get("/", self.index)

        self._runner = web.AppRunner(app, access_log=logger)
        await self._runner.setup()
        try:
            self._site = web.TCPSite(self._runner, WEB_HOST, WEB_PORT)
            await self._site.start()
        except Exception:
            await self._runner.cleanup()
            self._runner = None
            self._site = None
            raise
        logger.info("HLL Frontline web app listening on http://%s:%s", WEB_HOST, WEB_PORT)

    async def stop(self) -> None:
        if self._runner:
            await self._runner.cleanup()
            self._runner = None
            self._site = None

    @web.middleware
    async def _security_headers(self, request: web.Request, handler):
        response = await handler(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        turnstile_script = " https://challenges.cloudflare.com" if self._turnstile_enabled else ""
        turnstile_frame = "frame-src https://challenges.cloudflare.com; " if self._turnstile_enabled else ""
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; base-uri 'none'; frame-ancestors 'none'; form-action 'self'; "
            f"script-src 'self'{turnstile_script}; {turnstile_frame}connect-src 'self'; img-src 'self' data:; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com"
        )
        if request.secure or request.headers.get("X-Forwarded-Proto", "").casefold() == "https":
            response.headers["Strict-Transport-Security"] = "max-age=31536000"
        if request.path.startswith(("/api/", "/exports/")) or request.path == "/login":
            response.headers["Cache-Control"] = "no-store"
        return response

    @web.middleware
    async def _pin_auth(self, request: web.Request, handler):
        if self._is_sensitive_path(request.raw_path.split("?", 1)[0]):
            return web.Response(text="Not found.", status=404, content_type="text/plain")
        if request.path in {"/login", "/api/health"} or request.path.startswith("/assets/"):
            return await handler(request)
        if self._valid_session(request.cookies.get(SESSION_COOKIE, "")):
            return await handler(request)
        if request.path.startswith("/api/"):
            return web.json_response({"error": "Authentication required", "login_url": "/login"}, status=401)
        next_url = quote(self._safe_next(str(request.rel_url)), safe="")
        return web.HTTPSeeOther(location=f"/login?next={next_url}")

    @staticmethod
    def _is_sensitive_path(value: str) -> bool:
        path = str(value or "")
        # Decode twice to catch common encoded-dot probes without interpreting
        # the path as a filesystem location.
        for _ in range(2):
            decoded = unquote(path)
            if decoded == path:
                break
            path = decoded
        segments = [segment.casefold() for segment in path.replace("\\", "/").split("/") if segment]
        if any(segment.startswith(".") for segment in segments):
            return True
        sensitive_names = {
            "config.php",
            "credentials.json",
            "docker-compose.yml",
            "id_rsa",
            "secrets.json",
            "wp-config.php",
        }
        return any(segment in sensitive_names for segment in segments)

    @staticmethod
    def _safe_next(value: str | None) -> str:
        candidate = str(value or "/")
        return candidate if candidate.startswith("/") and not candidate.startswith("//") else "/"

    def _new_session(self, claimed_name: str) -> str:
        now = int(time.time())
        self._sessions = {
            token: session
            for token, session in self._sessions.items()
            if session.expires_at >= now
        }
        token = secrets.token_urlsafe(32)
        self._sessions[token] = WebSession(
            expires_at=now + SESSION_SECONDS,
            claimed_name=claimed_name,
        )
        return token

    def _valid_session(self, token: str) -> bool:
        session = self._sessions.get(token)
        if session is None:
            return False
        if session.expires_at < int(time.time()):
            self._sessions.pop(token, None)
            return False
        return True

    @staticmethod
    def _normalise_login_name(value: object) -> str | None:
        raw = str(value or "")
        if len(raw) > LOGIN_NAME_MAX_LENGTH or any(not character.isprintable() for character in raw):
            return None
        name = " ".join(raw.split())
        return name if name and len(name) <= LOGIN_NAME_MAX_LENGTH else None

    @staticmethod
    def _operator_login_notice(claimed_name: str) -> None:
        logging.getLogger().warning("%s has logged into your website!", claimed_name)

    @staticmethod
    def _client_key(request: web.Request) -> str:
        forwarded = (
            request.headers.get("CF-Connecting-IP")
            or request.headers.get("X-Real-IP")
            or request.headers.get("X-Forwarded-For", "").split(",", 1)[0].strip()
        )
        return forwarded or request.remote or "unknown"

    def _recent_failures(self, request: web.Request) -> list[float]:
        key = self._client_key(request)
        cutoff = time.monotonic() - LOGIN_WINDOW_SECONDS
        recent = [attempt for attempt in self._login_failures.get(key, []) if attempt >= cutoff]
        if recent:
            self._login_failures[key] = recent
        else:
            self._login_failures.pop(key, None)
        return recent

    @property
    def _turnstile_enabled(self) -> bool:
        return bool(self._turnstile_site_key and self._turnstile_secret_key)

    @staticmethod
    def _valid_turnstile_result(result: object, expected_hostname: str) -> bool:
        if not isinstance(result, dict) or result.get("success") is not True:
            return False
        hostname = str(result.get("hostname") or "").casefold()
        action = str(result.get("action") or "")
        return hmac.compare_digest(hostname, expected_hostname.casefold()) and action == "login"

    async def _verify_turnstile(self, request: web.Request, token: str) -> bool:
        if not self._turnstile_enabled or not token or len(token) > 2048:
            return False
        payload = {
            "secret": self._turnstile_secret_key,
            "response": token,
            "remoteip": self._client_key(request),
        }
        try:
            async with ClientSession(timeout=ClientTimeout(total=10)) as session:
                async with session.post(TURNSTILE_VERIFY_URL, data=payload) as response:
                    if response.status != 200:
                        logger.warning("Turnstile Siteverify returned HTTP %s", response.status)
                        return False
                    result = await response.json(content_type=None)
        except Exception:
            logger.warning("Turnstile Siteverify request failed", exc_info=True)
            return False
        return self._valid_turnstile_result(result, self._turnstile_hostname)

    async def login_page(self, request: web.Request) -> web.Response:
        if self._valid_session(request.cookies.get(SESSION_COOKIE, "")):
            return web.HTTPSeeOther(location=self._safe_next(request.query.get("next")))
        return self._login_document(
            next_url=self._safe_next(request.query.get("next")),
            error_code=str(request.query.get("error") or ""),
        )

    async def login(self, request: web.Request) -> web.Response:
        next_url = self._safe_next(request.query.get("next"))
        failures = self._recent_failures(request)
        if len(failures) >= LOGIN_MAX_FAILURES:
            return web.Response(
                text="Too many login attempts. Try again in 15 minutes.",
                status=429,
                content_type="text/plain",
                headers={"Retry-After": str(LOGIN_WINDOW_SECONDS)},
            )

        form = await request.post()
        next_url = self._safe_next(str(form.get("next") or next_url))
        claimed_name = self._normalise_login_name(form.get("name"))
        if claimed_name is None:
            return web.HTTPSeeOther(location=f"/login?error=name&next={quote(next_url, safe='')}")
        if self._turnstile_enabled:
            turnstile_token = str(form.get("cf-turnstile-response") or "")
            if not await self._verify_turnstile(request, turnstile_token):
                logger.warning(
                    "Rejected HLL Frontline login after failed Turnstile verification claimed_name=%r client_ip=%s",
                    claimed_name,
                    self._client_key(request),
                )
                return web.HTTPSeeOther(location=f"/login?error=turnstile&next={quote(next_url, safe='')}")
        supplied_pin = str(form.get("pin") or "")
        if not hmac.compare_digest(supplied_pin.encode("utf-8"), self._app_pin.encode("utf-8")):
            self._login_failures.setdefault(self._client_key(request), []).append(time.monotonic())
            logger.warning(
                "Rejected HLL Frontline PIN login claimed_name=%r client_ip=%s",
                claimed_name,
                self._client_key(request),
            )
            return web.HTTPSeeOther(location=f"/login?error=1&next={quote(next_url, safe='')}")

        client_ip = self._client_key(request)
        self._login_failures.pop(client_ip, None)
        logger.info("Successful HLL Frontline login claimed_name=%r client_ip=%s", claimed_name, client_ip)
        # The production service directs stderr to bot_error.log. Use the root
        # console logger for this explicit operator-facing notice while keeping
        # the detailed authentication audit in bot_web.log.
        self._operator_login_notice(claimed_name)
        response = web.HTTPSeeOther(location=next_url)
        response.set_cookie(
            SESSION_COOKIE,
            self._new_session(claimed_name),
            max_age=SESSION_SECONDS,
            httponly=True,
            secure=True,
            samesite="Strict",
            path="/",
        )
        return response

    async def logout(self, request: web.Request) -> web.Response:
        session = self._sessions.pop(request.cookies.get(SESSION_COOKIE, ""), None)
        if session is not None:
            logger.info(
                "HLL Frontline logout claimed_name=%r client_ip=%s",
                session.claimed_name,
                self._client_key(request),
            )
        response = web.HTTPSeeOther(location="/login")
        response.del_cookie(SESSION_COOKIE, path="/")
        return response

    def _login_document(self, *, next_url: str, error_code: str) -> web.Response:
        path = FRONTEND_DIR / "login.html"
        if not path.is_file():
            raise web.HTTPNotFound(text="Login frontend is missing.")
        document = path.read_text(encoding="utf-8")
        document = document.replace("{{NEXT}}", html.escape(next_url, quote=True))
        if self._turnstile_enabled:
            turnstile_head = (
                '<link rel="preconnect" href="https://challenges.cloudflare.com">\n'
                '  <script src="https://challenges.cloudflare.com/turnstile/v0/api.js" async defer></script>'
            )
            turnstile_widget = (
                f'<div class="cf-turnstile" data-sitekey="{html.escape(self._turnstile_site_key, quote=True)}" '
                'data-action="login" data-theme="dark"></div>'
            )
        else:
            turnstile_head = ""
            turnstile_widget = ""
        error_messages = {
            "1": "The PIN was not recognised.",
            "name": "Enter a valid name of no more than 80 characters.",
            "turnstile": "The security check was not completed. Please try again.",
        }
        document = document.replace("{{TURNSTILE_HEAD}}", turnstile_head)
        document = document.replace("{{TURNSTILE_WIDGET}}", turnstile_widget)
        document = document.replace("{{ERROR}}", error_messages.get(error_code, ""))
        return web.Response(text=document, content_type="text/html", charset="utf-8")

    async def health(self, _request: web.Request) -> web.Response:
        guild = self.bot.get_guild(MAIN_GUILD_ID)
        return web.json_response(
            {
                "status": "ok" if guild else "starting",
                "discord_ready": bool(self.bot.is_ready()),
                "guild_available": guild is not None,
            },
            status=200 if guild else 503,
        )

    def _allow_hllv_search(self, request: web.Request) -> bool:
        key = self._client_key(request)
        cutoff = time.monotonic() - HLLV_SEARCH_WINDOW_SECONDS
        recent = [attempt for attempt in self._hllv_searches.get(key, []) if attempt >= cutoff]
        if len(recent) >= HLLV_SEARCH_MAX_REQUESTS:
            self._hllv_searches[key] = recent
            return False
        recent.append(time.monotonic())
        self._hllv_searches[key] = recent
        return True

    async def hllv_search(self, request: web.Request) -> web.Response:
        if not self._allow_hllv_search(request):
            return web.json_response(
                {"error": "Too many searches. Try again in a minute."},
                status=429,
                headers={"Retry-After": str(HLLV_SEARCH_WINDOW_SECONDS)},
            )

        query = " ".join(str(request.query.get("q") or "").split())
        if not HLLV_SEARCH_MIN_LENGTH <= len(query) <= HLLV_SEARCH_MAX_LENGTH:
            return web.json_response(
                {"error": f"Enter between {HLLV_SEARCH_MIN_LENGTH} and {HLLV_SEARCH_MAX_LENGTH} characters."},
                status=400,
            )

        guild = self.bot.get_guild(MAIN_GUILD_ID)
        names = self.bot.get_cog("HLLVNames")
        if guild is None or names is None:
            raise web.HTTPServiceUnavailable(
                text='{"error":"The HLLV directory is unavailable"}',
                content_type="application/json",
            )

        records = names._search_records(guild, names._records(guild.id), query)[:20]
        results = []
        for user_id, hllv_name in records:
            member = guild.get_member(user_id)
            if member is None:
                continue
            results.append(
                {
                    "discord_name": str(member.display_name),
                    "hllv_name": str(hllv_name),
                }
            )
        return web.json_response({"query": query, "results": results})

    @staticmethod
    def _event_payload(guild) -> list[dict[str, Any]]:
        now = datetime.now(timezone.utc)
        events = []
        for event in getattr(guild, "scheduled_events", ()):
            status = str(getattr(getattr(event, "status", None), "name", getattr(event, "status", ""))).casefold()
            start = getattr(event, "start_time", None)
            if status not in {"scheduled", "active"} or not isinstance(start, datetime):
                continue
            if start.tzinfo is None:
                start = start.replace(tzinfo=timezone.utc)
            if status != "active" and start < now:
                continue
            end = getattr(event, "end_time", None)
            if isinstance(end, datetime) and end.tzinfo is None:
                end = end.replace(tzinfo=timezone.utc)
            events.append(
                {
                    "name": str(getattr(event, "name", "Scheduled event")),
                    "start_time": start.astimezone(timezone.utc).isoformat(),
                    "end_time": end.astimezone(timezone.utc).isoformat() if isinstance(end, datetime) else None,
                    "location": str(getattr(event, "location", "") or ""),
                    "url": str(getattr(event, "url", "") or ""),
                    "status": status,
                    "interested": int(getattr(event, "user_count", 0) or 0),
                }
            )
        return sorted(events, key=lambda item: item["start_time"])[:25]

    @staticmethod
    def _war_diary_payload(cog) -> dict[str, Any]:
        if cog is None:
            return {"summary": {"played": 0, "wins": 0, "losses": 0, "draws": 0}, "opponents": [], "recent": []}

        matches: list[dict[str, Any]] = []
        opponents: dict[str, dict[str, Any]] = {}
        for raw in cog._get_match_records():
            score_match = re.fullmatch(r"\s*(\d+)\s*[-:]\s*(\d+)\s*", str(raw.get("result") or ""))
            if score_match is None:
                continue
            home_score, away_score = (int(value) for value in score_match.groups())
            outcome = "win" if home_score > away_score else "loss" if home_score < away_score else "draw"
            opponent = " ".join(str(raw.get("opponent_clan_name") or "Unknown clan").split())
            opponent_key = opponent.casefold()
            record = opponents.setdefault(
                opponent_key,
                {"name": opponent, "played": 0, "wins": 0, "losses": 0, "draws": 0},
            )
            record["played"] += 1
            record[{"win": "wins", "loss": "losses", "draw": "draws"}[outcome]] += 1
            matches.append(
                {
                    "opponent": opponent,
                    "date": str(raw.get("match_date") or ""),
                    "map": str(raw.get("map_name") or "Unknown"),
                    "score": f"{home_score}-{away_score}",
                    "outcome": outcome,
                }
            )

        def date_key(item: dict[str, Any]) -> datetime:
            for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(item["date"], date_format)
                except ValueError:
                    continue
            return datetime.min

        matches.sort(key=date_key, reverse=True)
        opponent_rows = sorted(opponents.values(), key=lambda row: (-row["played"], row["name"].casefold()))
        wins = sum(match["outcome"] == "win" for match in matches)
        losses = sum(match["outcome"] == "loss" for match in matches)
        draws = sum(match["outcome"] == "draw" for match in matches)
        return {
            "summary": {"played": len(matches), "wins": wins, "losses": losses, "draws": draws},
            "opponents": opponent_rows,
            "recent": matches[:12],
        }

    @staticmethod
    def _botr_payload(guild) -> dict[str, Any]:
        role = guild.get_role(RAT_OF_THE_WEEK_ROLE_ID)
        if role is None:
            role = next(
                (candidate for candidate in getattr(guild, "roles", ()) if candidate.id == RAT_OF_THE_WEEK_ROLE_ID),
                None,
            )
        holders = sorted(
            (str(member.display_name) for member in getattr(role, "members", ()) if not getattr(member, "bot", False)),
            key=str.casefold,
        )
        return {"role_id": str(RAT_OF_THE_WEEK_ROLE_ID), "holders": holders}

    @staticmethod
    def _read_leaderboards(guild) -> list[dict[str, Any]]:
        from cogs.HLLArmLeaderboard import DB_FILE as ARM_DB_FILE, STATS_ARM, format_seconds_as_hhmmss, is_life_stat
        from cogs.HLLInfLeaderboard import DB_FILE as INF_DB_FILE, STATS

        def member_name(user_id: object) -> str:
            try:
                member = guild.get_member(int(user_id))
            except (TypeError, ValueError):
                member = None
            return str(member.display_name) if member is not None else "Former member"

        groups: list[dict[str, Any]] = []
        sources = (
            (
                "Infantry & Recon",
                Path(INF_DB_FILE),
                "user_id",
                STATS,
                "SELECT user_id, MAX(value) AS best FROM submissions "
                "WHERE stat = ? AND proof_verified = 1 "
                "GROUP BY user_id ORDER BY best DESC, user_id ASC LIMIT 3",
            ),
            (
                "Armour",
                Path(ARM_DB_FILE),
                "crew_key",
                STATS_ARM,
                "SELECT crew_key, MAX(value) AS best FROM submissions_arm "
                "WHERE stat = ? AND proof_verified = 1 "
                "GROUP BY crew_key ORDER BY best DESC, crew_key ASC LIMIT 3",
            ),
        )
        for title, path, owner_column, stats, query in sources:
            category = {"title": title, "records": []}
            if not path.is_file():
                groups.append(category)
                continue
            try:
                with sqlite3.connect(path) as database:
                    for stat in stats:
                        rows = database.execute(query, (stat,)).fetchall()
                        leaders = []
                        for owner, value in rows:
                            if owner_column == "crew_key":
                                names = [member_name(user_id) for user_id in str(owner).split(",") if user_id]
                                name = ", ".join(names) or "Former crew"
                                display_value = format_seconds_as_hhmmss(value) if is_life_stat(stat) else str(value)
                            else:
                                name = member_name(owner)
                                display_value = str(value)
                            leaders.append({"name": name, "value": display_value})
                        if leaders:
                            category["records"].append({"stat": stat, "leaders": leaders})
            except sqlite3.Error:
                logger.warning("Could not read %s website leaderboard", title, exc_info=True)
            groups.append(category)
        return groups

    @staticmethod
    def _coerce_server_data(value: object) -> dict[str, Any]:
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
            except json.JSONDecodeError:
                return {}
            return parsed if isinstance(parsed, dict) else {}
        return {}

    @classmethod
    def _normalise_server_status(cls, raw: dict[str, Any], fallback_name: str) -> dict[str, Any]:
        payload = cls._coerce_server_data(raw.get("data"))
        team1 = raw.get("team1") if isinstance(raw.get("team1"), dict) else {}
        team2 = raw.get("team2") if isinstance(raw.get("team2"), dict) else {}

        def first(*keys: str) -> str:
            for key in keys:
                value = payload.get(key)
                if value not in (None, ""):
                    return str(value)
            return ""

        current_map = first("currentMap", "current_map", "map", "mapName", "map_name", "server.map.name")
        game_mode = first("currentGameMode", "current_game_mode", "gameMode", "gamemode", "server.map.gamemode")
        players = 0
        for team in (team1, team2):
            try:
                players += int(team.get("playerCount") or 0)
            except (TypeError, ValueError):
                pass
        try:
            remaining = int(raw.get("matchTimeRemainingSeconds") or 0)
        except (TypeError, ValueError):
            remaining = 0
        return {
            "available": True,
            "name": first("serverName", "server_name", "server", "name", "server.name") or fallback_name,
            "map": " ".join(part for part in (current_map, game_mode) if part) or "Unknown",
            "players": players,
            "time_remaining_seconds": max(0, remaining),
            "next_map": str(raw.get("nextMap") or ""),
            "updated_at": str(raw.get("timestamp") or datetime.now(timezone.utc).isoformat()),
        }

    async def _server_status_payload(self) -> list[dict[str, Any]]:
        cached_at, cached = self._server_status_cache
        if time.monotonic() - cached_at < SERVER_STATUS_CACHE_SECONDS:
            return cached

        async with self._server_status_lock:
            cached_at, cached = self._server_status_cache
            if time.monotonic() - cached_at < SERVER_STATUS_CACHE_SECONDS:
                return cached

            from config.hll_API_config import get_hll_backend_status
            from hll_API_backend import get_hll_backend_client

            async def fetch(alias: str, label: str) -> dict[str, Any] | None:
                status = get_hll_backend_status(alias)
                if not status.get("server_id"):
                    return None
                try:
                    raw = await asyncio.wait_for(
                        get_hll_backend_client(alias).get_mapvote_game_state(),
                        timeout=12,
                    )
                    if not isinstance(raw, dict):
                        raise RuntimeError("empty server response")
                    return self._normalise_server_status(raw, label)
                except Exception as exc:
                    logger.warning("Website server status unavailable for %s: %s", alias, exc)
                    return {
                        "available": False,
                        "name": label,
                        "map": "Unknown",
                        "players": None,
                        "time_remaining_seconds": None,
                        "next_map": "",
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }

            results = await asyncio.gather(
                fetch("main", "7DR Public Server 1"),
                fetch("server_2", "7DR Public Server 2"),
            )
            payload = [result for result in results if result is not None]
            self._server_status_cache = (time.monotonic(), payload)
            return payload

    async def dashboard(self, _request: web.Request) -> web.Response:
        guild = self.bot.get_guild(MAIN_GUILD_ID)
        if guild is None:
            raise web.HTTPServiceUnavailable(
                text='{"error":"Discord data is not ready yet"}',
                content_type="application/json",
            )

        rollcall = self.bot.get_cog("RollCallCog")
        trainee = self.bot.get_cog("MultiTraineeTracker")
        if rollcall is None or trainee is None:
            raise web.HTTPServiceUnavailable(
                text='{"error":"Dashboard data providers are unavailable"}',
                content_type="application/json",
            )

        server_status_task = asyncio.create_task(self._server_status_payload())
        leaderboards_task = asyncio.create_task(asyncio.to_thread(self._read_leaderboards, guild))

        # The roll-call cog serialises workbook mutations with this lock. Reading under
        # the same lock prevents the API seeing a half-written weekly refresh.
        async with rollcall._lock:
            rollcalls = await asyncio.to_thread(self._rollcall_payload, rollcall, guild)
        trainees = self._trainee_payload(trainee, guild)
        server_status, leaderboards = await asyncio.gather(server_status_task, leaderboards_task)
        war_diary = self.bot.get_cog("WarDiaryCog")

        return web.json_response(
            {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "guild": {"name": guild.name},
                "external_links": EXTERNAL_LINKS,
                "server_status": server_status,
                "events": self._event_payload(guild),
                "war_diary": self._war_diary_payload(war_diary),
                "leaderboards": leaderboards,
                "botr": self._botr_payload(guild),
                "rollcalls": rollcalls,
                "trainee_tracks": trainees,
            }
        )

    def _rollcall_payload(self, cog, guild) -> list[dict[str, Any]]:
        workbook = cog._load_or_create_workbook()
        result: list[dict[str, Any]] = []

        for cfg in cog.ROLLCALLS if hasattr(cog, "ROLLCALLS") else ():
            # Kept for compatibility if configuration is moved onto the cog later.
            result.append(self._one_rollcall(cog, guild, workbook, cfg))

        if result:
            return result

        # Current cogs keep configuration at module scope.
        from cogs.rollcall import ROLLCALLS

        return [self._one_rollcall(cog, guild, workbook, cfg) for cfg in ROLLCALLS]

    def _one_rollcall(self, cog, guild, workbook, cfg) -> dict[str, Any]:
        ws = cog._get_or_create_sheet(workbook, cfg)
        state = cog._rc_state(cfg.key)
        week = state.get("current_week") or cog._week_label(cog._rollcall_date_for_now())
        status = cog._get_week_status(ws, week)
        expected = cog._expected_members(guild, cfg)
        expected_ids = {member.id for member in expected}

        def display_name(user_id: int, stored: str = "") -> str:
            member = guild.get_member(user_id)
            return member.display_name if member else stored or str(user_id)

        workbook_members: dict[int, str] = {}
        for row in range(2, ws.max_row + 1):
            try:
                user_id = int(str(ws.cell(row=row, column=1).value))
            except (TypeError, ValueError):
                continue
            workbook_members[user_id] = str(ws.cell(row=row, column=2).value or "")

        current_people = []
        for member in expected:
            value = status.get(member.id, "")
            current_people.append(
                {
                    "name": member.display_name,
                    "status": self._normalise_rollcall_status(value),
                }
            )

        headers = cog._sheet_headers(ws)
        week_headers = [header for header in headers[2:] if cog._parse_week_header(header)]
        history = []
        for header in reversed(week_headers[-12:]):
            values = cog._get_week_status(ws, header)
            active_values = [values.get(user_id, "") for user_id in expected_ids]
            history.append(
                {
                    "week": header,
                    "attending": sum(value == "✅" for value in active_values),
                    "partial": sum(value == "🅾️" for value in active_values),
                    "missing": sum(value not in ("✅", "🅾️") for value in active_values),
                }
            )

        report_members = []
        report_user_ids = list(workbook_members)
        report_user_ids.extend(user_id for user_id in expected_ids if user_id not in workbook_members)
        week_statuses = {header: cog._get_week_status(ws, header) for header in week_headers}
        for user_id in report_user_ids:
            member = guild.get_member(user_id)
            is_former = cog._is_former_member(member)
            active, flags = self._rollcall_member_state(user_id, expected_ids, member, is_former)
            rank, rank_order = self._member_rank(None if is_former else member)
            attendance = {
                header: week_statuses[header].get(user_id, "")
                for header in week_headers
            }
            report_members.append(
                {
                    "name": display_name(user_id, workbook_members.get(user_id, "")),
                    "rank": rank,
                    "rank_order": rank_order,
                    "flags": flags,
                    "active": active,
                    "missed_streak": self._missed_rollcall_streak(attendance, week_headers),
                    "attendance": attendance,
                }
            )

        departed_count = sum(not member["active"] for member in report_members)
        attending = sum(person["status"] == "attending" for person in current_people)

        return {
            "key": cfg.key,
            "title": cfg.title,
            "week": week,
            "locked": cog._is_rollcall_locked(state),
            "summary": {
                "total": len(current_people),
                "attending": attending,
                "missing": len(current_people) - attending,
                "rate": round((attending / len(current_people)) * 100) if current_people else 0,
            },
            "members": current_people,
            "history": history,
            "report_columns": week_headers,
            "report_members": report_members,
            "departed_count": departed_count,
        }

    @staticmethod
    def _rollcall_member_state(user_id: int, expected_ids: set[int], member, is_former: bool) -> tuple[bool, list[str]]:
        """Classify current role members separately from retained attendance history."""
        if member is None:
            return False, ["LEFT"]
        if is_former:
            return False, ["FORMER"]
        if user_id not in expected_ids:
            return False, ["ARCHIVED"]
        return True, []

    @staticmethod
    def _member_rank(member) -> tuple[str, int]:
        if member is None:
            return "Former member", len(DEFAULT_RANK_ORDER) + 1

        display_name = " ".join(str(getattr(member, "display_name", "") or "").strip().split())
        if "#" in display_name:
            display_name = display_name.split("#", 1)[0].strip()
        display_name_folded = display_name.casefold()

        # Clan ranks are normally prefixes in Discord display names. Prefer the
        # longest matching variant so "Lt Gen" is not mistaken for "Lt".
        display_matches: list[tuple[int, int, str]] = []
        for order, (code, variants) in enumerate(DEFAULT_RANK_ORDER):
            for variant in variants:
                prefix = variant.casefold().strip()
                if (
                    display_name_folded == prefix
                    or display_name_folded.startswith(prefix + " ")
                    or display_name_folded.startswith(prefix + ".")
                ):
                    display_matches.append((len(prefix), order, code))
        if display_matches:
            _length, order, code = max(display_matches, key=lambda match: match[0])
            return code, order

        # Retain role matching as a fallback for members whose nickname omits rank.
        def normalise(value: str) -> str:
            return re.sub(r"[^a-z0-9]", "", str(value).casefold())

        member_roles = {normalise(role.name) for role in getattr(member, "roles", [])}
        for order, (code, variants) in enumerate(DEFAULT_RANK_ORDER):
            accepted = {normalise(code), *(normalise(variant) for variant in variants)}
            if member_roles & accepted:
                return code, order
        return "Unranked", len(DEFAULT_RANK_ORDER)

    @staticmethod
    def _normalise_rollcall_status(value: str) -> str:
        if value == "✅":
            return "attending"
        if value == "🅾️":
            return "other-rollcall"
        return "missing"

    @staticmethod
    def _missed_rollcall_streak(attendance: dict[str, str], week_headers: list[str]) -> int:
        """Count consecutive explicit misses, starting with the newest roll call."""
        streak = 0
        for header in reversed(week_headers):
            if attendance.get(header) != "❌":
                break
            streak += 1
        return streak

    def _trainee_payload(self, cog, guild) -> list[dict[str, Any]]:
        from cogs.multi_trainee_tracker import BEHIND_AFTER_DAYS, TRACKS

        now = datetime.now(timezone.utc)
        result = []
        for cfg in TRACKS:
            rows = cog._collect_rows(guild, cfg)
            members = []
            for row in rows:
                joined = row["join_date"]
                if joined.tzinfo is None:
                    joined = joined.replace(tzinfo=timezone.utc)
                days = max(0, (now - joined).days)
                behind = joined < now - timedelta(days=BEHIND_AFTER_DAYS)
                members.append(
                    {
                        "name": row["display_name"],
                        "username": row["username"],
                        "joined": joined.date().isoformat(),
                        "review_due": row["plus_14"].date().isoformat(),
                        "days": days,
                        "behind": behind,
                        "checks": row["checks"],
                    }
                )
            result.append(
                {
                    "key": cfg.key,
                    "title": cfg.title,
                    "behind_after_days": BEHIND_AFTER_DAYS,
                    "summary": {
                        "total": len(members),
                        "behind": sum(member["behind"] for member in members),
                        "current": sum(not member["behind"] for member in members),
                    },
                    "check_labels": [label for label, _ in cfg.check_roles],
                    "members": members,
                }
            )
        return result

    async def asset(self, request: web.Request) -> web.StreamResponse:
        filename = request.match_info["filename"]
        if filename == "emblem_7dr.png":
            path = CLAN_EMBLEM_PATH
        elif filename in {"app.css", "app.js", "report.js"}:
            path = FRONTEND_DIR / filename
        else:
            raise web.HTTPNotFound()
        if not path.is_file():
            raise web.HTTPNotFound()
        return web.FileResponse(path)

    async def index(self, _request: web.Request) -> web.StreamResponse:
        path = FRONTEND_DIR / "index.html"
        if not path.is_file():
            raise web.HTTPNotFound(text="Frontend assets are missing.")
        return web.FileResponse(path)

    async def report_page(self, request: web.Request) -> web.StreamResponse:
        key = request.match_info["key"]
        if request.path.startswith("/rollcalls/"):
            from cogs.rollcall import ROLLCALLS

            valid_keys = {cfg.key for cfg in ROLLCALLS}
        else:
            from cogs.multi_trainee_tracker import TRACKS

            valid_keys = {cfg.key for cfg in TRACKS}
        if key not in valid_keys:
            raise web.HTTPNotFound()

        path = FRONTEND_DIR / "report.html"
        if not path.is_file():
            raise web.HTTPNotFound(text="Report frontend is missing.")
        return web.FileResponse(path)

    def _require_guild_and_cog(self, cog_name: str):
        guild = self.bot.get_guild(MAIN_GUILD_ID)
        cog = self.bot.get_cog(cog_name)
        if guild is None or cog is None:
            raise web.HTTPServiceUnavailable(text="Report data is not ready yet.")
        return guild, cog

    @staticmethod
    def _rollcall_config(key: str):
        from cogs.rollcall import ROLLCALLS

        config = next((cfg for cfg in ROLLCALLS if cfg.key == key), None)
        if config is None:
            raise web.HTTPNotFound()
        return config

    @staticmethod
    def _trainee_config(key: str):
        from cogs.multi_trainee_tracker import TRACKS

        config = next((cfg for cfg in TRACKS if cfg.key == key), None)
        if config is None:
            raise web.HTTPNotFound()
        return config

    async def rollcall_html_export(self, request: web.Request) -> web.Response:
        guild, cog = self._require_guild_and_cog("RollCallCog")
        config = self._rollcall_config(request.match_info["key"])
        async with cog._lock:
            workbook = await asyncio.to_thread(cog._load_or_create_workbook)
            worksheet = cog._get_or_create_sheet(workbook, config)
            state = cog._rc_state(config.key)
            week = state.get("current_week") or cog._week_label(cog._rollcall_date_for_now())
            document = cog._render_html(guild, worksheet, config, highlight_week=week)
        return web.Response(text=document, content_type="text/html", charset="utf-8")

    async def rollcall_excel_export(self, request: web.Request) -> web.Response:
        _guild, cog = self._require_guild_and_cog("RollCallCog")
        config = self._rollcall_config(request.match_info["key"])
        async with cog._lock:
            workbook = await asyncio.to_thread(cog._load_or_create_workbook)
            cog._get_or_create_sheet(workbook, config)
            output = io.BytesIO()
            await asyncio.to_thread(workbook.save, output)
        return self._excel_response(output.getvalue(), "rollcall.xlsx")

    async def trainee_html_export(self, request: web.Request) -> web.Response:
        guild, cog = self._require_guild_and_cog("MultiTraineeTracker")
        config = self._trainee_config(request.match_info["key"])
        rows = cog._collect_rows(guild, config)
        document = cog._render_html(config, rows)
        return web.Response(text=document, content_type="text/html", charset="utf-8")

    async def trainee_excel_export(self, request: web.Request) -> web.Response:
        guild, cog = self._require_guild_and_cog("MultiTraineeTracker")
        config = self._trainee_config(request.match_info["key"])
        rows = cog._collect_rows(guild, config)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = config.title[:31]
        headers = ["Name", "Username", "Join Date", "+14 Days"] + [label for label, _ in config.check_roles]
        worksheet.append(headers)
        for row in rows:
            worksheet.append(
                [
                    safe_spreadsheet_value(row["display_name"]),
                    safe_spreadsheet_value(row["username"]),
                    row["join_date"].date(),
                    row["plus_14"].date(),
                    *("Yes" if row["checks"].get(label) else "No" for label, _ in config.check_roles),
                ]
            )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column in worksheet.columns:
            letter = column[0].column_letter
            worksheet.column_dimensions[letter].width = min(
                42,
                max(12, max(len(str(cell.value or "")) for cell in column) + 2),
            )

        output = io.BytesIO()
        await asyncio.to_thread(workbook.save, output)
        return self._excel_response(output.getvalue(), f"{config.key}_trainee_tracker.xlsx")

    @staticmethod
    def _excel_response(content: bytes, filename: str) -> web.Response:
        return web.Response(
            body=content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


async def setup(bot) -> None:
    service = FrontlineWeb(bot)
    await service.start()
    bot.frontline_web = service


async def teardown(bot) -> None:
    service = getattr(bot, "frontline_web", None)
    if service:
        await service.stop()
        del bot.frontline_web
